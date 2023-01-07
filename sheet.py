import xlwt
import xlrd
import openpyxl
from openpyxl.utils import get_column_letter
from xlutils.copy import copy
import os
import datetime
import xlsxwriter

st_name = 'Ayushman'
def mark_present(st_name):

	names = os.listdir('output/')
	print(names)

	sub = 'SAMPLE'
	
	if not os.path.exists('attendance/' + sub + '.xlsx'):
		count = 2
		wb = openpyxl.Workbook()
		wb = openpyxl.load_workbook('attendance/' + sub + '.xlsx')
		print("Creating Spreadsheet with Title: " + sub)
		sheet = wb.create_sheet("testsheet") 
		sheet = wb.active
		for i in names:
		    sheet.append([count, i])
		    count += 1
		wb.save('attendance/' + sub + '.xlsx') 

	wb_obj = openpyxl.Workbook()
	wb_obj = openpyxl.load_workbook('attendance/' + sub + '.xlsx')
	#wb_obj = copy(rb)
	sheet = wb_obj['Sheet1']
	sheet['B1'].value = str(datetime.datetime.now())
	# sheet.append([1,str(datetime.datetime.now())])

	count = 2
	for i in names:
		var='B'+str(count)
		if i in st_name:
			sheet[var].value = 'Pr'
		else:
			sheet[var].value = 'Ab'
		var1='A'+str(count)
		sheet[var1].value = i
		count += 1

#	for i in names:
#	    if i in st_name:
#            sheet.write(count, 1, 'P')
#	    else:
#            sheet.write(count, 1, 'A')
#	    sheet.write(count, 0, i)
#	    count += 1

	wb_obj.save('attendance/' + sub + '.xlsx')


mark_present(st_name)
